#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import io
import json
import os
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple
from urllib.parse import parse_qs, urlparse

from sqlalchemy import MetaData, Table, and_, create_engine, select
from sqlalchemy.engine import Connection, Engine
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.sql import ClauseElement

try:
    from sqlalchemy.dialects.postgresql import insert as pg_insert
except Exception:  # pragma: no cover
    pg_insert = None

try:
    import requests  # type: ignore
except ImportError:  # pragma: no cover
    requests = None  # type: ignore


JsonDict = Dict[str, Any]


class ImportConfigError(ValueError):
    pass


@dataclass
class RowIssue:
    row_number: int
    reason: str
    row: Dict[str, Any]


def _load_json(path: str) -> JsonDict:
    with open(path, "r", encoding="utf-8") as handle:
        data = json.load(handle)
    if not isinstance(data, dict):
        raise ImportConfigError("Config must be a JSON object.")
    return data


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_url_host(value: Any) -> str:
    raw = _clean_text(value)
    if not raw:
        return ""
    try:
        parsed = urlparse(raw if "://" in raw else f"https://{raw}")
        host = (parsed.hostname or "").strip().lower()
        return host
    except Exception:
        trimmed = raw.lower()
        if "://" in trimmed:
            trimmed = trimmed.split("://", 1)[1]
        return trimmed.split("/", 1)[0].strip().lower()


def _normalize_url_canonical(value: Any) -> str:
    raw = _clean_text(value)
    if not raw:
        return ""
    try:
        parsed = urlparse(raw if "://" in raw else f"https://{raw}")
        host = (parsed.hostname or "").strip().lower()
        path = (parsed.path or "").rstrip("/")
        if path:
            return f"{host}{path}"
        return host
    except Exception:
        return raw.strip().lower().rstrip("/")


def _apply_transform(value: Any, transform: Optional[str]) -> Any:
    if not transform:
        return value
    key = transform.strip().lower()
    if key == "trim":
        return _clean_text(value)
    if key == "lower":
        return _clean_text(value).lower()
    if key == "upper":
        return _clean_text(value).upper()
    if key == "url_host":
        return _normalize_url_host(value)
    if key == "url_canonical":
        return _normalize_url_canonical(value)
    if key == "int":
        text = _clean_text(value)
        return int(text) if text else None
    if key == "bool":
        text = _clean_text(value).lower()
        if text in {"1", "true", "yes", "y", "on"}:
            return True
        if text in {"0", "false", "no", "n", "off", ""}:
            return False
        raise ImportConfigError(f"Unsupported bool value '{value}'.")
    if key == "null_if_blank":
        text = _clean_text(value)
        return text or None
    raise ImportConfigError(f"Unknown transform '{transform}'.")


def _coerce_nulls(value: Any, null_values: List[str]) -> Any:
    if value is None:
        return None
    if isinstance(value, str) and value.strip() in null_values:
        return None
    return value


def _parse_google_sheet_csv_url(input_cfg: JsonDict) -> str:
    raw_url = _clean_text(input_cfg.get("url"))
    sheet_id = _clean_text(input_cfg.get("sheet_id"))
    gid = _clean_text(input_cfg.get("gid")) or "0"
    if raw_url:
        parsed = urlparse(raw_url)
        if "docs.google.com" in (parsed.netloc or "") and "/spreadsheets/" in (parsed.path or ""):
            parts = parsed.path.split("/d/")
            if len(parts) > 1:
                sheet_id = parts[1].split("/", 1)[0]
            query = parse_qs(parsed.query or "")
            if query.get("gid"):
                gid = query["gid"][0]
        elif raw_url.endswith(".csv"):
            return raw_url
    if not sheet_id:
        raise ImportConfigError("Google Sheets input requires 'url' or 'sheet_id'.")
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"


def _read_csv_bytes(data: bytes, *, encoding: str) -> List[Dict[str, Any]]:
    text = data.decode(encoding, errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ImportConfigError("Input file has no header row.")
    return [dict(row) for row in reader]


def _read_xlsx_rows(path: str, *, sheet_name: Optional[str]) -> List[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:  # pragma: no cover
        raise ImportConfigError("Excel import requires 'openpyxl'. Install it in the environment.") from exc

    workbook = load_workbook(path, data_only=True, read_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active
    rows_iter = worksheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration as exc:
        raise ImportConfigError("Excel sheet is empty.") from exc
    headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
    if not any(headers):
        raise ImportConfigError("Excel sheet header row is empty.")
    result: List[Dict[str, Any]] = []
    for row in rows_iter:
        payload: Dict[str, Any] = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            payload[header] = row[idx] if idx < len(row) else None
        result.append(payload)
    return result


def read_rows_from_input(input_cfg: JsonDict) -> List[Dict[str, Any]]:
    input_type = _clean_text(input_cfg.get("type") or "csv").lower()
    encoding = _clean_text(input_cfg.get("encoding") or "utf-8")

    if input_type == "csv":
        path = _clean_text(input_cfg.get("path"))
        if not path:
            raise ImportConfigError("CSV input requires 'input.path'.")
        with open(path, "rb") as handle:
            return _read_csv_bytes(handle.read(), encoding=encoding)

    if input_type == "xlsx":
        path = _clean_text(input_cfg.get("path"))
        if not path:
            raise ImportConfigError("XLSX input requires 'input.path'.")
        return _read_xlsx_rows(path, sheet_name=_clean_text(input_cfg.get("sheet_name")) or None)

    if input_type in {"google_sheet", "gsheet", "google-sheets"}:
        if requests is None:
            raise ImportConfigError("Google Sheets input requires the 'requests' package.")
        url = _parse_google_sheet_csv_url(input_cfg)
        timeout = int(input_cfg.get("timeout_seconds") or 30)
        response = requests.get(url, timeout=timeout)
        response.raise_for_status()
        return _read_csv_bytes(response.content, encoding=encoding)

    raise ImportConfigError(f"Unsupported input.type '{input_type}'. Use csv/xlsx/google_sheet.")


def _build_engine(config: JsonDict) -> Engine:
    db_url = _clean_text(config.get("database_url"))
    if not db_url:
        env_name = _clean_text(config.get("database_url_env") or "DATABASE_URL")
        db_url = _clean_text(os.getenv(env_name))
    if not db_url:
        raise ImportConfigError("Database URL missing. Set config.database_url or config.database_url_env.")
    return create_engine(db_url, future=True)


def _reflect_table(engine: Engine, table_name: str) -> Table:
    metadata = MetaData()
    return Table(table_name, metadata, autoload_with=engine)


def _literal_spec_value(spec: JsonDict) -> Any:
    if "value" in spec:
        return spec["value"]
    if "default" in spec:
        return spec["default"]
    return None


def _row_blank(row: Dict[str, Any]) -> bool:
    for value in row.values():
        if _clean_text(value):
            return False
    return True


class LookupResolver:
    def __init__(self, engine: Engine, config: JsonDict):
        self.engine = engine
        self.config = config
        self._cache: Optional[Dict[str, Any]] = None

    def build(self) -> Dict[str, Any]:
        if self._cache is not None:
            return self._cache
        table = _reflect_table(self.engine, _clean_text(self.config["table"]))
        match_column = _clean_text(self.config["match_column"])
        return_column = _clean_text(self.config["return_column"])
        if match_column not in table.c:
            raise ImportConfigError(f"Lookup match_column '{match_column}' not found in table '{table.name}'.")
        if return_column not in table.c:
            raise ImportConfigError(f"Lookup return_column '{return_column}' not found in table '{table.name}'.")

        normalize = _clean_text(self.config.get("lookup_normalize") or "")
        where_cfg = self.config.get("where")
        clauses: List[ClauseElement] = []
        if isinstance(where_cfg, dict):
            for col_name, expected in where_cfg.items():
                if col_name not in table.c:
                    raise ImportConfigError(f"Lookup where column '{col_name}' not found in table '{table.name}'.")
                clauses.append(table.c[col_name] == expected)

        query = select(table.c[match_column], table.c[return_column])
        if clauses:
            query = query.where(and_(*clauses))

        result: Dict[str, Any] = {}
        with self.engine.connect() as conn:
            for match_value, return_value in conn.execute(query):
                key_raw = match_value
                key = _apply_transform(key_raw, normalize or None)
                if key in ("", None):
                    continue
                result[str(key)] = return_value
        self._cache = result
        return result

    def resolve(self, raw_value: Any) -> Any:
        mapping = self.build()
        normalize = _clean_text(self.config.get("lookup_normalize") or "")
        key = _apply_transform(raw_value, normalize or None)
        if key in ("", None):
            return None
        return mapping.get(str(key))


def _resolve_column_value(
    source_row: Dict[str, Any],
    spec: Any,
    *,
    lookup_cache: Dict[str, LookupResolver],
    null_values: List[str],
) -> Any:
    if isinstance(spec, str):
        value = source_row.get(spec)
        return _coerce_nulls(value, null_values)
    if not isinstance(spec, dict):
        return spec

    if "lookup" in spec:
        lookup_spec = spec["lookup"]
        if not isinstance(lookup_spec, dict):
            raise ImportConfigError("lookup spec must be an object.")
        source_col = _clean_text(lookup_spec.get("source_column"))
        if not source_col:
            raise ImportConfigError("lookup.source_column is required.")
        cache_key = json.dumps(lookup_spec, sort_keys=True)
        resolver = lookup_cache.get(cache_key)
        if resolver is None:
            resolver = LookupResolver(engine=lookup_cache["__engine__"].engine, config=lookup_spec)  # type: ignore[attr-defined]
            lookup_cache[cache_key] = resolver
        raw_value = source_row.get(source_col)
        resolved = resolver.resolve(_coerce_nulls(raw_value, null_values))
        if resolved is None:
            if "default" in spec:
                return spec["default"]
            return None
        value = resolved
    elif "source" in spec:
        value = source_row.get(spec["source"])
    elif "value" in spec or "default" in spec:
        value = _literal_spec_value(spec)
    else:
        value = None

    value = _coerce_nulls(value, null_values)
    if "transform" in spec:
        value = _apply_transform(value, spec.get("transform"))
    if "default" in spec and (value is None or (isinstance(value, str) and value == "")):
        value = spec["default"]
    return value


class _EngineRef:
    def __init__(self, engine: Engine):
        self.engine = engine


def prepare_rows(
    raw_rows: List[Dict[str, Any]],
    *,
    config: JsonDict,
    engine: Engine,
    target_table: Table,
) -> Tuple[List[Dict[str, Any]], List[RowIssue]]:
    mapping_cfg = config.get("column_map")
    if not isinstance(mapping_cfg, dict) or not mapping_cfg:
        raise ImportConfigError("config.column_map must be a non-empty object.")

    null_values = [str(v) for v in (config.get("null_values") or ["", "NULL", "null", "None"])]
    skip_blank_rows = bool(config.get("skip_blank_rows", True))
    required_target_columns = set(config.get("required_columns") or [])

    prepared: List[Dict[str, Any]] = []
    issues: List[RowIssue] = []
    lookup_cache: Dict[str, Any] = {"__engine__": _EngineRef(engine)}

    for idx, raw_row in enumerate(raw_rows, start=2):  # assumes header row is line 1
        row = dict(raw_row)
        if skip_blank_rows and _row_blank(row):
            continue
        try:
            target_row: Dict[str, Any] = {}
            for target_column, spec in mapping_cfg.items():
                if target_column not in target_table.c:
                    raise ImportConfigError(f"Target column '{target_column}' does not exist in table '{target_table.name}'.")
                value = _resolve_column_value(row, spec, lookup_cache=lookup_cache, null_values=null_values)
                spec_obj = spec if isinstance(spec, dict) else {}
                if spec_obj.get("required") and (value is None or value == ""):
                    raise ValueError(f"Required value missing for target column '{target_column}'.")
                if "lookup" in spec_obj and value is None and spec_obj.get("required", False):
                    source_col = _clean_text(spec_obj["lookup"].get("source_column"))
                    raise ValueError(f"Lookup failed for target column '{target_column}' from source column '{source_col}'.")
                target_row[target_column] = value

            for col_name in required_target_columns:
                if target_row.get(col_name) in (None, ""):
                    raise ValueError(f"Required column '{col_name}' is missing.")

            prepared.append(target_row)
        except Exception as exc:
            issues.append(RowIssue(row_number=idx, reason=str(exc), row=row))
    return prepared, issues


def _chunked(items: List[Dict[str, Any]], size: int) -> Iterable[List[Dict[str, Any]]]:
    for i in range(0, len(items), size):
        yield items[i:i + size]


def _dialect_supports_pg_upsert(engine: Engine) -> bool:
    return engine.dialect.name == "postgresql" and pg_insert is not None


def _manual_upsert(
    conn: Connection,
    table: Table,
    rows: List[Dict[str, Any]],
    *,
    match_columns: List[str],
    update_columns: List[str],
) -> Tuple[int, int]:
    inserted = 0
    updated = 0
    for row in rows:
        where_clauses = [table.c[col] == row.get(col) for col in match_columns]
        existing = conn.execute(select(table).where(and_(*where_clauses)).limit(1)).mappings().first()
        if existing:
            payload = {col: row.get(col) for col in update_columns if col in row}
            if payload:
                conn.execute(table.update().where(and_(*where_clauses)).values(**payload))
            updated += 1
        else:
            conn.execute(table.insert().values(**row))
            inserted += 1
    return inserted, updated


def apply_upsert(
    engine: Engine,
    table: Table,
    rows: List[Dict[str, Any]],
    *,
    match_columns: List[str],
    update_columns: Optional[List[str]],
    batch_size: int = 500,
    dry_run: bool = False,
) -> Tuple[int, int]:
    if not rows:
        return 0, 0
    for col in match_columns:
        if col not in table.c:
            raise ImportConfigError(f"Match column '{col}' does not exist in target table '{table.name}'.")

    resolved_update_columns = update_columns or [col for col in rows[0].keys() if col not in set(match_columns)]
    for col in resolved_update_columns:
        if col not in table.c:
            raise ImportConfigError(f"Update column '{col}' does not exist in target table '{table.name}'.")

    if dry_run:
        return 0, 0

    inserted = 0
    updated = 0
    with engine.begin() as conn:
        if _dialect_supports_pg_upsert(engine):
            for chunk in _chunked(rows, batch_size):
                stmt = pg_insert(table).values(chunk)
                update_payload = {col: getattr(stmt.excluded, col) for col in resolved_update_columns}
                stmt = stmt.on_conflict_do_update(index_elements=match_columns, set_=update_payload)
                result = conn.execute(stmt)
                # Postgres rowcount for upsert combines inserts + updates.
                if result.rowcount and result.rowcount > 0:
                    updated += int(result.rowcount)
        else:
            for chunk in _chunked(rows, batch_size):
                i, u = _manual_upsert(conn, table, chunk, match_columns=match_columns, update_columns=resolved_update_columns)
                inserted += i
                updated += u
    return inserted, updated


def _print_issue_summary(issues: List[RowIssue], limit: int = 20) -> None:
    if not issues:
        return
    print(f"\nSkipped rows: {len(issues)}")
    for issue in issues[:limit]:
        print(f"  row {issue.row_number}: {issue.reason}")
    if len(issues) > limit:
        print(f"  ... {len(issues) - limit} more")


def _write_issues_jsonl(path: str, issues: List[RowIssue]) -> None:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    with target.open("w", encoding="utf-8") as handle:
        for issue in issues:
            handle.write(
                json.dumps(
                    {
                        "row_number": issue.row_number,
                        "reason": issue.reason,
                        "row": issue.row,
                    },
                    ensure_ascii=False,
                )
                + "\n"
            )


def run_import(config: JsonDict, *, dry_run: bool) -> int:
    engine = _build_engine(config)
    table_name = _clean_text(config.get("table"))
    if not table_name:
        raise ImportConfigError("config.table is required.")
    target_table = _reflect_table(engine, table_name)

    raw_rows = read_rows_from_input(config.get("input") or {})
    prepared_rows, issues = prepare_rows(raw_rows, config=config, engine=engine, target_table=target_table)

    match_columns = [str(c) for c in (config.get("match_columns") or [])]
    if not match_columns:
        raise ImportConfigError("config.match_columns is required for upsert.")
    update_columns = [str(c) for c in (config.get("update_columns") or [])] or None

    batch_size = int(config.get("batch_size") or 500)
    inserted, updated = apply_upsert(
        engine,
        target_table,
        prepared_rows,
        match_columns=match_columns,
        update_columns=update_columns,
        batch_size=batch_size,
        dry_run=dry_run,
    )

    print(f"Target table: {table_name}")
    print(f"Input rows: {len(raw_rows)}")
    print(f"Prepared rows: {len(prepared_rows)}")
    print(f"Skipped rows: {len(issues)}")
    if dry_run:
        print("Mode: DRY RUN (no database changes applied)")
    else:
        if _dialect_supports_pg_upsert(engine):
            print(f"Upserted rows (insert/update combined): {updated}")
        else:
            print(f"Inserted rows: {inserted}")
            print(f"Updated rows: {updated}")

    _print_issue_summary(issues)
    issues_out = _clean_text(config.get("issues_output_jsonl"))
    if issues_out and issues:
        _write_issues_jsonl(issues_out, issues)
        print(f"Issues written to: {issues_out}")

    return 0 if not issues else (0 if bool(config.get("allow_issues", True)) else 2)


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="General tabular data importer (CSV/XLSX/Google Sheets) with configurable DB upsert + lookup mapping."
    )
    parser.add_argument("--config", required=True, help="Path to JSON config file.")
    parser.add_argument("--dry-run", action="store_true", help="Validate + preview without writing to DB.")
    return parser


def main() -> int:
    parser = build_arg_parser()
    args = parser.parse_args()
    try:
        config = _load_json(args.config)
        return run_import(config, dry_run=bool(args.dry_run))
    except (ImportConfigError, SQLAlchemyError, OSError) as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1
    except Exception as exc:
        if requests is not None and isinstance(exc, requests.RequestException):
            print(f"Error: {exc}", file=sys.stderr)
            return 1
        print(f"Error: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
